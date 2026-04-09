import os, re

# ── Paths ─────────────────────────────────────────────────────────────────────
# Adjust these three paths for your own machine/project.
#
# ANALYSIS_DIR  — folder containing one subfolder per subject named
#                 {S##}_AnalysisExplanation/, each holding a text file:
#                   Gaze2_{S##}_T1_AnalysisSheet.txt
#                 This is the EEGLAB manual analysis sheet exported from the
#                 original Latinus et al. processing pipeline.
#                 Example layout:
#                   analysis/
#                     S01_AnalysisExplanation/
#                       Gaze2_S01_T1_AnalysisSheet.txt
#                     S03_AnalysisExplanation/
#                       Gaze2_S03_T1_AnalysisSheet.txt
#                     ...
#
# BL_DIR        — folder containing bad-epoch outputs downloaded from Brainlife,
#                 one subfolder per subject+pass named {S##}_first/ or {S##}_second/.
#                 Each subfolder must contain events.tsv with 0-based epoch indices
#                 as comma-separated values (Brainlife app-detect-bad-epochs output).
#                 Example layout:
#                   brainlife_bad_epochs/
#                     S01_first/
#                       events.tsv      ← "15,22,29,..."
#                     S01_second/
#                       events.tsv
#                     S03_first/
#                       events.tsv
#                     ...
#
# OUT_DIR       — where bad_epochs_summary.tsv, bad_epochs_detail.tsv, and
#                 bad_epochs.xlsx will be written (can be same as ANALYSIS_DIR).
ANALYSIS_DIR = "/path/to/your/analysis_dir"          # folder containing S##_AnalysisExplanation/ subfolders
BL_DIR       = "/path/to/your/brainlife_bad_epochs"  # folder containing S##_first/ and S##_second/ subfolders
OUT_DIR      = "/path/to/your/output_dir"            # where TSV and Excel outputs will be written

# ── Collect all subject IDs from BOTH sources ─────────────────────────────────
# A subject is included if it has EITHER a Brainlife folder OR an analysis sheet.
# (Previously only analysis folders were scanned — this missed S24 whose sheet is empty.)

all_subjects = set()

# From analysis folders (S01_AnalysisExplanation/ etc.)
for folder in os.listdir(ANALYSIS_DIR):
    m = re.match(r'(S\d+)_AnalysisExplanation', folder)
    if m:
        all_subjects.add(m.group(1))

# From Brainlife folders (S24_first/, S24_second/ etc.)
for folder in os.listdir(BL_DIR):
    m = re.match(r'(S\d+)_(first|second)', folder)
    if m:
        all_subjects.add(m.group(1))

all_subjects = sorted(all_subjects)


# ── Helper: read one Brainlife bad-epoch file ─────────────────────────────────
# Brainlife stores epoch indices as 0-based comma-separated values, e.g.:
#   15,22,29,38,...
# Returns sorted list of ints (0-based). Caller adds +1 for EEGLAB 1-based comparison.
# Returns None if the file doesn't exist (subject has no second pass, etc.).

def read_bl_epochs(subject, pass_name):
    path = os.path.join(BL_DIR, f"{subject}_{pass_name}", "events.tsv")
    if not os.path.exists(path):
        return None
    with open(path) as f:
        content = f.read().strip().rstrip(',')   # Brainlife sometimes adds trailing comma
    if not content:
        return []
    return sorted([int(x) for x in content.split(',') if x.strip()])


# ── Helper: parse one analysis sheet text file ────────────────────────────────
# Returns a dict. Called only when the sheet file exists and is non-empty.
# Returns None if the file is empty or unreadable.

def parse_sheet(text):
    notes = []

    # ── Isolate pre-ICA section ───────────────────────────────────────────────
    # re.split with (?m)^ means "start of any line" (multiline mode).
    # maxsplit=1 → split at FIRST "ICA file:" line only → [before_ICA, after_ICA].
    # [0] takes the "before ICA" part.
    ica_split    = re.split(r'(?m)^ICA file:', text, maxsplit=1)
    pre_ica_text = ica_split[0]

    # ── Find all BTmark blocks in the pre-ICA section ────────────────────────
    # re.findall returns ALL non-overlapping matches as a list of tuples.
    # Each tuple is (count_string, indices_string), e.g. ('50', '16 23 30 ...')
    #
    #   Bad Trials marked epochs file:  — literal anchor text
    #   .*?                             — any chars, non-greedy (as few as possible)
    #   \n                              — newline after the file path line
    #   Total No\. of rejected epochs:  — literal (\.  escapes the dot)
    #   \s*(\d+)\s*\n                   — group 1: the count number
    #   Indices of rejected epochs:     — literal
    #   \s*([\d\s]+)                    — group 2: space-separated index list
    btmark_blocks = re.findall(
        r'Bad Trials marked epochs file:.*?\nTotal No\. of rejected epochs:\s*(\d+)\s*\n'
        r'Indices of rejected epochs:\s*([\d\s]+)',
        pre_ica_text
    )

    # Convert each block's index string ("16   23   30 ...") to a sorted list of ints
    pre_ica_marks = [sorted([int(x) for x in idx.split()]) for _, idx in btmark_blocks]

    if len(btmark_blocks) == 0:
        notes.append("NO BTmark recorded")
    elif len(btmark_blocks) == 2:
        notes.append("2 BTmark iterations — rejected once using final list")

    # The LAST BTmark entry is the one actually applied
    pre_ica_final = pre_ica_marks[-1] if pre_ica_marks else None

    # ── Trials remaining after pre-ICA rejection ─────────────────────────────
    # re.search finds the FIRST match anywhere in the string.
    # re.DOTALL makes . match newlines too (needed to span multiple lines).
    # .group(1) returns the first captured () group.
    btrej_match   = re.search(
        r'Bad trials rejected file:.*?\nTotal No\. of trials left:\s*(\d+)',
        pre_ica_text, re.DOTALL
    )
    pre_ica_count = int(btrej_match.group(1)) if btrej_match else None

    # ── Post-ICA artifact rejection block (searches full text) ───────────────
    # Same pattern logic as BTmark above.
    # Searches full `text` (not pre_ica_text) — this section appears after ICA.
    artifact_match = re.search(
        r'Artifact marked epochs file:.*?\nTotal No\. of rejected epochs:\s*(\d+)\s*\n'
        r'Indices of rejected epochs:\s*([\d\s]+)',
        text
    )
    post_ica_indices = sorted([int(x) for x in artifact_match.group(2).split()]) if artifact_match else None
    if not artifact_match:
        notes.append("No post-ICA artifact rejection recorded")

    # Trials remaining after artifact rejection (final clean trial count)
    final_match   = re.search(
        r'Artifact rejected file:.*?\nTotal No\. of trials left:\s*(\d+)',
        text, re.DOTALL
    )
    post_ica_count = int(final_match.group(1)) if final_match else None

    return {
        "pre_ica_marks":    pre_ica_marks,
        "pre_ica_final":    pre_ica_final,
        "pre_ica_count":    pre_ica_count,
        "post_ica_indices": post_ica_indices,
        "post_ica_count":   post_ica_count,
        "notes":            notes,
    }


# ── Parse each subject ────────────────────────────────────────────────────────
subjects = []

for sub in all_subjects:
    sheet_path = os.path.join(ANALYSIS_DIR, f"{sub}_AnalysisExplanation",
                              f"Gaze2_{sub}_T1_AnalysisSheet.txt")

    bl_first  = read_bl_epochs(sub, 'first')
    bl_second = read_bl_epochs(sub, 'second')

    # Skip entirely if NEITHER source has data
    has_sheet = os.path.exists(sheet_path) and os.path.getsize(sheet_path) > 0
    has_bl    = bl_first is not None or bl_second is not None
    if not has_sheet and not has_bl:
        continue

    if has_sheet:
        with open(sheet_path) as f:
            text = f.read()
        parsed = parse_sheet(text)
    else:
        # Fallback: no analysis sheet — fill with None, note it
        parsed = {
            "pre_ica_marks":    [],
            "pre_ica_final":    None,
            "pre_ica_count":    None,
            "post_ica_indices": None,
            "post_ica_count":   None,
            "notes":            ["no analysis sheet available"],
        }

    if not has_bl:
        parsed["notes"].append("no Brainlife bad epoch data")

    parsed['subject'] = sub
    subjects.append(parsed)


# ── Build output rows ─────────────────────────────────────────────────────────

summary_rows = ['\t'.join([
    'subject', 'pre_ica_n_passes', 'pre_ica_n_rejected', 'pre_ica_trials_left',
    'post_ica_n_rejected', 'post_ica_trials_left',
    'bl_first_n', 'bl_second_n', 'bl_first_meaning', 'bl_second_meaning', 'notes'
])]

detail_rows = ['\t'.join(['subject', 'stage', 'source', 'n_epochs', 'indices_1based'])]

for s in subjects:
    sub       = s['subject']
    bl_first  = read_bl_epochs(sub, 'first')
    bl_second = read_bl_epochs(sub, 'second')
    n_marks   = len(s['pre_ica_marks'])
    no_sheet  = 'no analysis sheet available' in s['notes']

    # Determine what Brainlife "first" and "second" actually mean for this subject.
    # Varies because Brainlife stores passes in the order they were logged:
    #   - No sheet:        unknown (can only describe by count)
    #   - S06 (0 BTmarks): first = post-ICA (no pre-ICA recorded)
    #   - S03 (2 BTmarks): first = intermediate mark, second = final mark (both pre-ICA)
    #   - Everyone else:   first = pre-ICA bad trials, second = post-ICA artifact
    if no_sheet:
        bl_first_meaning  = 'unknown (no analysis sheet)'
        bl_second_meaning = 'unknown (no analysis sheet)' if bl_second is not None else 'N/A'
    elif n_marks == 0:
        bl_first_meaning  = 'post-ICA artifact rejection'
        bl_second_meaning = 'N/A'
    elif n_marks == 2:
        bl_first_meaning  = 'pre-ICA BTmark iter1 (intermediate)'
        bl_second_meaning = 'pre-ICA BTmark iter2 (final applied)'
    else:
        bl_first_meaning  = 'pre-ICA bad trials'
        bl_second_meaning = 'post-ICA artifact rejection (re-indexed)'

    summary_rows.append('\t'.join([
        sub, str(n_marks),
        str(len(s['pre_ica_final'])) if s['pre_ica_final'] else '?',
        str(s['pre_ica_count'])      if s['pre_ica_count']  else '?',
        str(len(s['post_ica_indices'])) if s['post_ica_indices'] else '?',
        str(s['post_ica_count'])        if s['post_ica_count']   else '?',
        str(len(bl_first))  if bl_first  is not None else 'N/A',
        str(len(bl_second)) if bl_second is not None else 'N/A',
        bl_first_meaning, bl_second_meaning,
        ' | '.join(s['notes']) if s['notes'] else 'normal',
    ]))

    # Detail: sheet side
    for i, mark_list in enumerate(s['pre_ica_marks'], 1):
        detail_rows.append('\t'.join([
            sub, f'pre_ica_BTmark_iter{i}', 'sheet_eeglab',
            str(len(mark_list)), ' '.join(str(x) for x in mark_list)
        ]))
    if s['post_ica_indices']:
        detail_rows.append('\t'.join([
            sub, 'post_ica_artifact', 'sheet_eeglab',
            str(len(s['post_ica_indices'])), ' '.join(str(x) for x in s['post_ica_indices'])
        ]))

    # Detail: Brainlife side (0-based → +1 for EEGLAB 1-based comparison)
    for pass_name, bl_data in [('first', bl_first), ('second', bl_second)]:
        if bl_data is not None:
            bl_1based = [x + 1 for x in bl_data]
            detail_rows.append('\t'.join([
                sub, f'brainlife_{pass_name}', 'brainlife_plus1',
                str(len(bl_1based)), ' '.join(str(x) for x in bl_1based)
            ]))


# ── Write TSV output files ────────────────────────────────────────────────────
with open(os.path.join(OUT_DIR, 'bad_epochs_summary.tsv'), 'w') as f:
    f.write('\n'.join(summary_rows))
with open(os.path.join(OUT_DIR, 'bad_epochs_detail.tsv'), 'w') as f:
    f.write('\n'.join(detail_rows))


# ── Write fancy Excel file ────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

COLOURS = {
    'header':        'FF2C3E50',   # dark navy
    'header_font':   'FFFFFFFF',   # white
    'pre_ica_sheet': 'FFDCE6F1',   # light blue — sheet pre-ICA rows
    'post_ica_sheet':'FFDAEEF3',   # light teal — sheet post-ICA rows
    'brainlife':     'FFE2EFDA',   # light green — brainlife rows
    'subject_bg':    'FFFFF2CC',   # light yellow — merged subject cell
    'border':        'FFB8B8B8',   # grey border
}

thin   = Side(style='thin', color=COLOURS['border'])
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_cell(cell, bg=None, bold=False, wrap=False, align='left'):
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.font      = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border    = border


# ── Sheet 1: Detail ───────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Detail"

headers = ['Subject', 'Stage', 'Source', 'N Epochs', 'Indices (1-based, space separated)']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill      = PatternFill('solid', fgColor=COLOURS['header'])
    cell.font      = Font(bold=True, color=COLOURS['header_font'], size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = border
ws.row_dimensions[1].height = 20

data_rows = [r.split('\t') for r in detail_rows[1:]]
row_num   = 2
sub_start = {}
sub_end   = {}

for r in data_rows:
    sub, stage, source, n, indices = r
    if sub not in sub_start:
        sub_start[sub] = row_num
    sub_end[sub] = row_num

    if source == 'brainlife_plus1':
        bg = COLOURS['brainlife']
    elif 'post_ica' in stage:
        bg = COLOURS['post_ica_sheet']
    else:
        bg = COLOURS['pre_ica_sheet']

    cell = ws.cell(row=row_num, column=1, value=sub)
    style_cell(cell, bg=COLOURS['subject_bg'], bold=True, align='center')
    ws.cell(row=row_num, column=2, value=stage)
    style_cell(ws.cell(row=row_num, column=2), bg=bg)
    ws.cell(row=row_num, column=3, value=source)
    style_cell(ws.cell(row=row_num, column=3), bg=bg)
    ws.cell(row=row_num, column=4, value=int(n))
    style_cell(ws.cell(row=row_num, column=4), bg=bg, align='center')
    ws.cell(row=row_num, column=5, value=indices)
    style_cell(ws.cell(row=row_num, column=5), bg=bg, wrap=True)
    row_num += 1

# Merge subject column cells vertically (one tall cell per subject)
for sub, start in sub_start.items():
    end = sub_end[sub]
    if end > start:
        ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
    cell = ws.cell(row=start, column=1)
    cell.fill      = PatternFill('solid', fgColor=COLOURS['subject_bg'])
    cell.font      = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = border

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 80
ws.freeze_panes = 'A2'


# ── Sheet 2: Summary ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")

sum_headers = summary_rows[0].split('\t')
for col, h in enumerate(sum_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill      = PatternFill('solid', fgColor=COLOURS['header'])
    cell.font      = Font(bold=True, color=COLOURS['header_font'], size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = border
ws2.row_dimensions[1].height = 20

for row_i, row_str in enumerate(summary_rows[1:], 2):
    vals = row_str.split('\t')
    for col_i, val in enumerate(vals, 1):
        cell = ws2.cell(row=row_i, column=col_i, value=val)
        if col_i == 1:
            # Subject column: always yellow + bold
            style_cell(cell, bg=COLOURS['subject_bg'], bold=True, align='center')
        elif col_i == len(vals) and val != 'normal':
            # Notes column: orange if non-normal
            style_cell(cell, bg='FFFCE4D6', wrap=True)
        else:
            bg = COLOURS['pre_ica_sheet'] if row_i % 2 == 0 else 'FFFFFFFF'
            style_cell(cell, bg=bg)

sum_widths = [10, 14, 16, 18, 16, 18, 10, 11, 34, 38, 40]
for col_i, w in enumerate(sum_widths, 1):
    ws2.column_dimensions[get_column_letter(col_i)].width = w
ws2.freeze_panes = 'A2'


xlsx_path = os.path.join(OUT_DIR, 'bad_epochs.xlsx')
wb.save(xlsx_path)
print(f"Done — Excel: {xlsx_path}")
print("Subjects:", [s['subject'] for s in subjects])
